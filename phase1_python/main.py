#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
管理栄養士業務システム - メインエントリポイント
Phase1: Python + CSV
"""

import argparse
import csv
import os
import sys
from datetime import datetime
from calendar import monthrange
from typing import Optional

try:
    from openpyxl import Workbook
except ImportError:
    Workbook = None  # type: ignore

# プロジェクトルート（phase1_python の親）
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(PROJECT_ROOT, "data")

# 栄養価の目標値（保育園3〜5歳児）
TARGET_ENERGY = 650
TARGET_PROTEIN = 25
TARGET_FAT = 20
TARGET_CARBOHYDRATE = 85
TARGET_SALT = 3.0


def load_csv(filename: str) -> list[dict]:
    """CSVファイルを読み込み、リストの辞書として返す"""
    filepath = os.path.join(DATA_DIR, filename)
    if not os.path.exists(filepath):
        print(f"エラー: {filename} が存在しません")
        sys.exit(1)

    with open(filepath, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
    # 空行を除外
    return [r for r in rows if any(r.values())]


def validate_reference_integrity(
    ingredients: list[dict], dishes: list[dict], recipe: list[dict]
) -> tuple[bool, list[str]]:
    """
    参照整合性をチェック。戻り値: (OKか, エラーメッセージのリスト)
    テスト用の純粋関数。
    """
    errors: list[str] = []
    ingredient_ids = {int(i["id"]) for i in ingredients}
    dish_ids = {int(d["id"]) for d in dishes}

    for row in recipe:
        dish_id = int(row["dish_id"])
        ing_id = int(row["ingredient_id"])
        if dish_id not in dish_ids:
            errors.append(f"recipe の dish_id={dish_id} が dishes に存在しません")
        if ing_id not in ingredient_ids:
            errors.append(f"recipe の ingredient_id={ing_id} が ingredients に存在しません")
    return (len(errors) == 0, errors)


def check_reference_integrity(
    ingredients: list[dict], dishes: list[dict], recipe: list[dict]
) -> None:
    """参照整合性をチェック。違反があればエラー終了"""
    ok, errors = validate_reference_integrity(ingredients, dishes, recipe)
    if not ok:
        for e in errors:
            print(f"エラー: {e}")
        sys.exit(1)
    print("参照整合性: OK")


def get_ingredient_by_id(ingredients: list[dict], ing_id: int) -> Optional[dict]:
    """食材IDから食材データを取得"""
    for i in ingredients:
        if int(i["id"]) == ing_id:
            return i
    return None


def calc_dish_nutrition(
    dish_id: int,
    recipe: list[dict],
    ingredients: list[dict],
) -> dict[str, float]:
    """
    1料理の栄養価を計算
    戻り値: {"energy", "protein", "fat", "carbohydrate", "salt"}
    """
    result = {"energy": 0.0, "protein": 0.0, "fat": 0.0, "carbohydrate": 0.0, "salt": 0.0}
    nut_keys = ["energy", "protein", "fat", "carbohydrate", "salt"]

    for row in recipe:
        if int(row["dish_id"]) != dish_id:
            continue
        ing_id = int(row["ingredient_id"])
        amount = float(row["amount"])
        ing = get_ingredient_by_id(ingredients, ing_id)
        if ing is None:
            continue
        for key in nut_keys:
            result[key] += float(ing[key]) * amount / 100

    # 丸め適用（要件定義書準拠）
    result["energy"] = round(result["energy"], 1)
    result["protein"] = round(result["protein"], 1)
    result["fat"] = round(result["fat"], 1)
    result["carbohydrate"] = round(result["carbohydrate"], 1)
    result["salt"] = round(result["salt"], 2)
    return result


def calc_menu_nutrition(
    menu_row: dict,
    recipe: list[dict],
    ingredients: list[dict],
) -> dict[str, float]:
    """献立全体の栄養価を計算"""
    dish_ids = [
        int(menu_row["staple_id"]),
        int(menu_row["main_id"]),
        int(menu_row["side_id"]),
        int(menu_row["soup_id"]),
        int(menu_row["dessert_id"]),
    ]
    result = {"energy": 0.0, "protein": 0.0, "fat": 0.0, "carbohydrate": 0.0, "salt": 0.0}
    nut_keys = ["energy", "protein", "fat", "carbohydrate", "salt"]

    for dish_id in dish_ids:
        dish_nut = calc_dish_nutrition(dish_id, recipe, ingredients)
        for key in nut_keys:
            result[key] += dish_nut[key]

    # 丸め適用（合計後）
    result["energy"] = round(result["energy"], 1)
    result["protein"] = round(result["protein"], 1)
    result["fat"] = round(result["fat"], 1)
    result["carbohydrate"] = round(result["carbohydrate"], 1)
    result["salt"] = round(result["salt"], 2)
    return result


def cmd_calc(
    args: argparse.Namespace,
    ingredients: list[dict],
    dishes: list[dict],
    recipe: list[dict],
    menus: list[dict],
) -> None:
    """calc --date の処理"""
    date_str = args.date
    menu_row = None
    for m in menus:
        if m.get("date") == date_str:
            menu_row = m
            break
    if menu_row is None:
        print(f"エラー: {date_str} の献立が登録されていません")
        sys.exit(1)

    nut = calc_menu_nutrition(menu_row, recipe, ingredients)

    # 目標値との比較表示
    targets = {
        "energy": TARGET_ENERGY,
        "protein": TARGET_PROTEIN,
        "fat": TARGET_FAT,
        "carbohydrate": TARGET_CARBOHYDRATE,
        "salt": TARGET_SALT,
    }
    labels = {
        "energy": "エネルギー",
        "protein": "たんぱく質",
        "fat": "脂質",
        "carbohydrate": "炭水化物",
        "salt": "食塩相当量",
    }
    units = {"energy": "kcal", "protein": "g", "fat": "g", "carbohydrate": "g", "salt": "g"}

    print(f"献立の栄養価（{date_str}）")
    print("-" * 40)
    for key in ["energy", "protein", "fat", "carbohydrate", "salt"]:
        v = nut[key]
        t = targets[key]
        diff = round(v - t, 1) if key != "salt" else round(v - t, 2)
        diff_str = f"+{diff}" if diff >= 0 else str(diff)
        if key in ["energy", "protein"]:
            print(f"{labels[key]}：{v} {units[key]}（目標{t}{units[key]}：{diff_str}）")
        else:
            print(f"{labels[key]}：{v} {units[key]}")


MENU_CATEGORIES = [
    ("主食", "staple_id"),
    ("主菜", "main_id"),
    ("副菜", "side_id"),
    ("汁物", "soup_id"),
    ("デザート", "dessert_id"),
]


def get_dishes_by_category(dishes: list[dict]) -> dict[str, list[dict]]:
    """献立候補をカテゴリ別に分類"""
    result: dict[str, list[dict]] = {}
    for d in dishes:
        cat = d["menu_category"]
        if cat not in result:
            result[cat] = []
        result[cat].append(d)
    return result


def cmd_create_menu(
    args: argparse.Namespace,
    dishes: list[dict],
    menus: list[dict],
) -> None:
    """create-menu --date の処理（対話式で献立を登録）"""
    date_str = args.date

    # 同日登録の禁止
    for m in menus:
        if m.get("date") == date_str:
            print(f"エラー: {date_str} は既に登録されています。同日登録は禁止です。")
            sys.exit(1)

    dishes_by_cat = get_dishes_by_category(dishes)
    selected: dict[str, int] = {}

    print(f"献立登録（{date_str}）")
    print("-" * 40)

    for cat_label, col_name in MENU_CATEGORIES:
        cat_dishes = dishes_by_cat.get(cat_label, [])
        if not cat_dishes:
            print(f"エラー: {cat_label} の献立候補がありません")
            sys.exit(1)

        print(f"\n【{cat_label}】")
        for i, d in enumerate(cat_dishes, 1):
            print(f"  {i}. {d['name']} (id:{d['id']})")

        while True:
            try:
                choice = input(f"  {cat_label}を選択 (1-{len(cat_dishes)}): ").strip()
                idx = int(choice)
                if 1 <= idx <= len(cat_dishes):
                    selected[col_name] = int(cat_dishes[idx - 1]["id"])
                    break
            except ValueError:
                pass
            print("  無効な入力です。正しい番号を入力してください。")

    # menus.csv に追記
    menus_path = os.path.join(DATA_DIR, "menus.csv")
    need_header = not os.path.exists(menus_path) or os.path.getsize(menus_path) == 0

    with open(menus_path, "a", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["date", "staple_id", "main_id", "side_id", "soup_id", "dessert_id", "note"],
            lineterminator="\n",
        )
        if need_header:
            writer.writeheader()
        row = {
            "date": date_str,
            "staple_id": selected["staple_id"],
            "main_id": selected["main_id"],
            "side_id": selected["side_id"],
            "soup_id": selected["soup_id"],
            "dessert_id": selected["dessert_id"],
            "note": "",
        }
        writer.writerow(row)

    print("-" * 40)
    print(f"献立を登録しました: {date_str}")


def aggregate_order(
    menus: list[dict],
    recipe: list[dict],
    start_date: str,
    end_date: str,
    people: int,
) -> dict[int, float]:
    """
    期間内の献立から食材使用量を集計。戻り値: ingredient_id -> 総使用量(g)
    同一食材は合算、人数倍率を適用。テスト用の純粋関数。
    """
    target_menus = [
        m for m in menus
        if m.get("date") and start_date <= m["date"] <= end_date
    ]

    amount_by_ing: dict[int, float] = {}
    for menu_row in target_menus:
        dish_ids = [
            int(menu_row["staple_id"]),
            int(menu_row["main_id"]),
            int(menu_row["side_id"]),
            int(menu_row["soup_id"]),
            int(menu_row["dessert_id"]),
        ]
        for dish_id in dish_ids:
            for row in recipe:
                if int(row["dish_id"]) != dish_id:
                    continue
                ing_id = int(row["ingredient_id"])
                amount = float(row["amount"])
                amount_by_ing[ing_id] = amount_by_ing.get(ing_id, 0) + amount

    return {
        ing_id: round(amt * people, 1) for ing_id, amt in amount_by_ing.items()
    }


def cmd_order(
    args: argparse.Namespace,
    ingredients: list[dict],
    recipe: list[dict],
    menus: list[dict],
) -> None:
    """order --start --end --people の処理（発注集計）"""
    start_date = args.start
    end_date = args.end
    people = args.people

    target_menus = [
        m for m in menus
        if m.get("date") and start_date <= m["date"] <= end_date
    ]
    if not target_menus:
        print(f"エラー: {start_date} 〜 {end_date} に献立が登録されていません")
        sys.exit(1)

    total_by_ing = aggregate_order(menus, recipe, start_date, end_date, people)
    ing_by_id = {int(i["id"]): i["name"] for i in ingredients}

    print(f"発注集計（{start_date} 〜 {end_date}、{people}人分）")
    print("-" * 40)
    print(f"{'食材名':<20} {'総使用量(g)':>12}")
    print("-" * 40)

    for ing_id in sorted(total_by_ing.keys()):
        name = ing_by_id.get(ing_id, f"不明(id:{ing_id})")
        total = total_by_ing[ing_id]
        print(f"{name:<20} {total:>12.1f}")


WEEKDAY_JA = ["月", "火", "水", "木", "金", "土", "日"]


def get_dish_by_id(dishes: list[dict], dish_id: int) -> Optional[dict]:
    """料理IDから料理データを取得"""
    for d in dishes:
        if int(d["id"]) == dish_id:
            return d
    return None


def cmd_export(
    args: argparse.Namespace,
    ingredients: list[dict],
    dishes: list[dict],
    recipe: list[dict],
    menus: list[dict],
) -> None:
    """export --month --output の処理（献立表Excel出力）"""
    if Workbook is None:
        print("エラー: openpyxl がインストールされていません。pip install openpyxl を実行してください。")
        sys.exit(1)

    month_str = args.month  # YYYY-MM
    output_path = args.output

    # 月の開始日・終了日
    try:
        year, month = map(int, month_str.split("-"))
        start_date = f"{year:04d}-{month:02d}-01"
        _, last_day = monthrange(year, month)
        end_date = f"{year:04d}-{month:02d}-{last_day:02d}"
    except ValueError:
        print(f"エラー: --month は YYYY-MM 形式で指定してください（例: 2026-04）")
        sys.exit(1)

    # 期間内の献立を取得（日付順）
    target_menus = [
        m for m in menus
        if m.get("date") and start_date <= m["date"] <= end_date
    ]
    target_menus.sort(key=lambda m: m["date"])

    if not target_menus:
        print(f"エラー: {month_str} に献立が登録されていません")
        sys.exit(1)

    # Excel作成
    wb = Workbook()
    ws = wb.active
    ws.title = month_str.replace("-", "年") + "月"

    # ヘッダー
    headers = ["日付", "曜日", "主食", "主菜", "副菜", "汁物", "デザート", "エネルギー"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # データ行
    for row_idx, menu_row in enumerate(target_menus, 2):
        date_str = menu_row["date"]
        try:
            dt = datetime.strptime(date_str, "%Y-%m-%d")
            weekday = WEEKDAY_JA[dt.weekday()]  # 0=月, 6=日
        except ValueError:
            weekday = ""

        # 料理名を取得
        staple = get_dish_by_id(dishes, int(menu_row["staple_id"]))
        main_d = get_dish_by_id(dishes, int(menu_row["main_id"]))
        side = get_dish_by_id(dishes, int(menu_row["side_id"]))
        soup = get_dish_by_id(dishes, int(menu_row["soup_id"]))
        dessert = get_dish_by_id(dishes, int(menu_row["dessert_id"]))

        nut = calc_menu_nutrition(menu_row, recipe, ingredients)

        ws.cell(row=row_idx, column=1, value=date_str)
        ws.cell(row=row_idx, column=2, value=weekday)
        ws.cell(row=row_idx, column=3, value=staple["name"] if staple else "")
        ws.cell(row=row_idx, column=4, value=main_d["name"] if main_d else "")
        ws.cell(row=row_idx, column=5, value=side["name"] if side else "")
        ws.cell(row=row_idx, column=6, value=soup["name"] if soup else "")
        ws.cell(row=row_idx, column=7, value=dessert["name"] if dessert else "")
        ws.cell(row=row_idx, column=8, value=nut["energy"])

    # 出力パス（相対パスの場合はプロジェクトルート基準）
    if not os.path.isabs(output_path):
        output_path = os.path.join(PROJECT_ROOT, output_path)

    wb.save(output_path)
    print(f"献立表を出力しました: {output_path}")


def run_step1_check() -> None:
    """ステップ1: データスキーマ固定の動作確認"""
    print("管理栄養士業務システム - ステップ1 動作確認")
    print("-" * 40)

    ingredients = load_csv("ingredients.csv")
    dishes = load_csv("dishes.csv")
    recipe = load_csv("recipe.csv")
    menus = load_csv("menus.csv")

    print(f"食材: {len(ingredients)} 件読み込みました")
    print(f"献立候補: {len(dishes)} 件読み込みました")
    print(f"レシピ: {len(recipe)} 件読み込みました")
    print(f"献立: {len(menus)} 件読み込みました")

    check_reference_integrity(ingredients, dishes, recipe)

    print("-" * 40)
    print("ステップ1 完了: データスキーマ固定 OK")


def main() -> None:
    parser = argparse.ArgumentParser(description="管理栄養士業務システム")
    subparsers = parser.add_subparsers(dest="command", help="コマンド")

    # calc コマンド
    calc_parser = subparsers.add_parser("calc", help="栄養価計算")
    calc_parser.add_argument("--date", required=True, help="日付 (YYYY-MM-DD)")
    calc_parser.set_defaults(command="calc")

    # create-menu コマンド
    create_parser = subparsers.add_parser("create-menu", help="献立登録")
    create_parser.add_argument("--date", required=True, help="日付 (YYYY-MM-DD)")
    create_parser.set_defaults(command="create-menu")

    # order コマンド
    order_parser = subparsers.add_parser("order", help="発注集計")
    order_parser.add_argument("--start", required=True, help="開始日 (YYYY-MM-DD)")
    order_parser.add_argument("--end", required=True, help="終了日 (YYYY-MM-DD)")
    order_parser.add_argument("--people", type=int, required=True, help="人数")
    order_parser.set_defaults(command="order")

    # export コマンド
    export_parser = subparsers.add_parser("export", help="献立表Excel出力")
    export_parser.add_argument("--month", required=True, help="対象月 (YYYY-MM)")
    export_parser.add_argument("--output", required=True, help="出力ファイル名 (.xlsx)")
    export_parser.set_defaults(command="export")

    args = parser.parse_args()

    if args.command == "calc":
        ingredients = load_csv("ingredients.csv")
        dishes = load_csv("dishes.csv")
        recipe = load_csv("recipe.csv")
        menus = load_csv("menus.csv")
        check_reference_integrity(ingredients, dishes, recipe)
        cmd_calc(args, ingredients, dishes, recipe, menus)
    elif args.command == "create-menu":
        dishes = load_csv("dishes.csv")
        menus = load_csv("menus.csv")
        cmd_create_menu(args, dishes, menus)
    elif args.command == "order":
        ingredients = load_csv("ingredients.csv")
        dishes = load_csv("dishes.csv")
        recipe = load_csv("recipe.csv")
        menus = load_csv("menus.csv")
        check_reference_integrity(ingredients, dishes, recipe)
        cmd_order(args, ingredients, recipe, menus)
    elif args.command == "export":
        ingredients = load_csv("ingredients.csv")
        dishes = load_csv("dishes.csv")
        recipe = load_csv("recipe.csv")
        menus = load_csv("menus.csv")
        check_reference_integrity(ingredients, dishes, recipe)
        cmd_export(args, ingredients, dishes, recipe, menus)
    else:
        run_step1_check()


if __name__ == "__main__":
    main()
